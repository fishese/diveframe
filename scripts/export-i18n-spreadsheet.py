"""Export app i18n keys in approximate on-screen order to an Excel workbook.

Helper columns (do not edit when translating):
  screenOrder — 1-based order as keys appear across walked UI screens
  fileOrder   — 1-based order in the English source definition (overlay first,
                then app-i18n/en.ts). Used to re-sort back to file order.
  file        — which source owns the key: lib/i18n.ts or lib/app-i18n
  screen      — UI section label where the key first appeared
"""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUT = Path(r"C:\Users\Selena\OneDrive\Documents\i18n-list.xlsx")

SCREEN_FILES: list[tuple[str, Path]] = [
    ("00 Global chrome", ROOT / "app" / "BetaNotice.tsx"),
    ("01 Home — top bar & overview", ROOT / "app" / "DiveFrameApp.tsx"),
    ("02 Import guide", ROOT / "app" / "components" / "ImportGuide.tsx"),
    ("03 Dive memos", ROOT / "app" / "memos" / "MemosApp.tsx"),
    ("04 Composer", ROOT / "app" / "compose" / "ComposerApp.tsx"),
    ("05 Settings", ROOT / "app" / "settings" / "SettingsApp.tsx"),
    ("06 About", ROOT / "app" / "about" / "AboutApp.tsx"),
    ("07 Android page", ROOT / "app" / "android" / "AndroidAppPage.tsx"),
    ("08 Shared components", ROOT / "app" / "components" / "AndroidAppLink.tsx"),
]

KEY_RE = re.compile(r"""\bt\(\s*["']([A-Za-z][A-Za-z0-9_]*)["']""")
PAIR_RE = re.compile(
    r'^  ([A-Za-z][A-Za-z0-9_]*)\s*:\s*("(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\')',
    re.M,
)


def unescape_js_string(raw: str) -> str:
    body = raw[1:-1]
    out: list[str] = []
    i = 0
    while i < len(body):
        if body[i] == "\\" and i + 1 < len(body):
            nxt = body[i + 1]
            mapping = {"n": "\n", "t": "\t", "\\": "\\", '"': '"', "'": "'"}
            out.append(mapping.get(nxt, nxt))
            i += 2
            continue
        out.append(body[i])
        i += 1
    return "".join(out)


def parse_object_block(text: str, marker: str) -> dict[str, str]:
    start = text.find(marker)
    if start < 0:
        return {}
    brace = text.find("{", start)
    if brace < 0:
        return {}
    depth = 0
    end = brace
    for i, ch in enumerate(text[brace:], start=brace):
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    block = text[brace : end + 1]
    return {m.group(1): unescape_js_string(m.group(2)) for m in PAIR_RE.finditer(block)}


def parse_locale_ordered(path: Path) -> tuple[dict[str, str], list[str]]:
    text = path.read_text(encoding="utf-8")
    ordered = [m.group(1) for m in PAIR_RE.finditer(text)]
    values = {m.group(1): unescape_js_string(m.group(2)) for m in PAIR_RE.finditer(text)}
    return values, ordered


def load_translations() -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, int],
]:
    """Returns en, ja, zh, key->file, key->fileOrder."""
    overlay = (ROOT / "lib" / "i18n.ts").read_text(encoding="utf-8")
    overlay_en = parse_object_block(overlay, "en:")
    overlay_ja = parse_object_block(overlay, "ja:")
    overlay_zh = parse_object_block(overlay, "zh-Hant:")
    # File order: overlay en keys first (definition order), then app-i18n/en.ts
    overlay_order = list(overlay_en.keys())
    app_en, app_order = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "en.ts")
    app_ja, _ = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "ja.ts")
    app_zh, _ = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "zh-Hant.ts")

    en = {**overlay_en, **app_en}
    ja = {**overlay_ja, **app_ja}
    zh = {**overlay_zh, **app_zh}

    key_file: dict[str, str] = {}
    file_order: dict[str, int] = {}
    idx = 1
    for key in overlay_order:
        key_file[key] = "lib/i18n.ts"
        file_order[key] = idx
        idx += 1
    for key in app_order:
        # app-i18n overrides overlay for the same key in runtime merge
        key_file[key] = "lib/app-i18n"
        file_order[key] = idx
        idx += 1
    # Any JA/ZH-only leftovers
    for key in set(en) | set(ja) | set(zh):
        if key not in file_order:
            key_file[key] = "lib/app-i18n"
            file_order[key] = idx
            idx += 1

    return en, ja, zh, key_file, file_order


def extract_keys(path: Path) -> list[str]:
    if not path.exists():
        return []
    text = path.read_text(encoding="utf-8")
    seen: set[str] = set()
    ordered: list[str] = []
    for match in KEY_RE.finditer(text):
        key = match.group(1)
        if key in seen:
            continue
        seen.add(key)
        ordered.append(key)
    return ordered


def main() -> None:
    en, ja, zh, key_file, file_order = load_translations()

    wb = Workbook()
    sheet = wb.active
    sheet.title = "All (screen order)"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D2731")
    helper_fill = PatternFill("solid", fgColor="1A3A44")
    section_fill = PatternFill("solid", fgColor="143946")
    section_font = Font(bold=True, color="8DEBD7")
    thin = Border(
        left=Side(style="thin", color="2A4A55"),
        right=Side(style="thin", color="2A4A55"),
        top=Side(style="thin", color="2A4A55"),
        bottom=Side(style="thin", color="2A4A55"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # Editable: key, EN, JA, ZH. Helpers: screenOrder, fileOrder, file, screen
    headers = [
        "key",
        "EN",
        "JA",
        "ZH",
        "screenOrder",
        "fileOrder",
        "file",
        "screen",
    ]
    sheet.append(headers)
    for col, name in enumerate(headers, start=1):
        cell = sheet.cell(1, col)
        cell.font = header_font
        cell.fill = helper_fill if col >= 5 else header_fill
        cell.alignment = Alignment(vertical="center")

    used: set[str] = set()
    screen_order = 0
    data_rows = 0

    for screen, path in SCREEN_FILES:
        sheet.append([f"— {screen} —", "", "", "", "", "", "", ""])
        r = sheet.max_row
        for col in range(1, 9):
            cell = sheet.cell(r, col)
            cell.fill = section_fill
            cell.font = section_font

        for key in extract_keys(path):
            if key in used:
                continue
            if key not in en and key not in ja and key not in zh:
                continue
            used.add(key)
            screen_order += 1
            data_rows += 1
            sheet.append(
                [
                    key,
                    en.get(key, ""),
                    ja.get(key, ""),
                    zh.get(key, ""),
                    screen_order,
                    file_order.get(key, ""),
                    key_file.get(key, ""),
                    screen,
                ]
            )
            r = sheet.max_row
            for col in range(1, 9):
                cell = sheet.cell(r, col)
                cell.alignment = wrap
                cell.border = thin

    leftover = sorted((set(en) | set(ja) | set(zh)) - used)
    if leftover:
        sheet.append(
            ["— Not referenced in walked UI files —", "", "", "", "", "", "", ""]
        )
        r = sheet.max_row
        for col in range(1, 9):
            cell = sheet.cell(r, col)
            cell.fill = section_fill
            cell.font = section_font
        for key in leftover:
            screen_order += 1
            data_rows += 1
            sheet.append(
                [
                    key,
                    en.get(key, ""),
                    ja.get(key, ""),
                    zh.get(key, ""),
                    screen_order,
                    file_order.get(key, ""),
                    key_file.get(key, ""),
                    "99 leftover",
                ]
            )
            r = sheet.max_row
            for col in range(1, 9):
                cell = sheet.cell(r, col)
                cell.alignment = wrap
                cell.border = thin

    widths = {
        "A": 34,
        "B": 48,
        "C": 48,
        "D": 48,
        "E": 12,
        "F": 10,
        "G": 14,
        "H": 30,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"

    # README sheet
    readme = wb.create_sheet("README", 0)
    readme["A1"] = "How to edit this workbook"
    readme["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "Edit only: key (do not rename), EN, JA, ZH.",
        "Leave helper columns alone: screenOrder, fileOrder, file, screen.",
        "",
        "screenOrder — on-screen / spreadsheet walk order (1, 2, 3…).",
        "fileOrder   — order in English source files (lib/i18n.ts overlay first,",
        "              then lib/app-i18n/en.ts). Sort by this column to match",
        "              the TypeScript files; not required to apply updates",
        "              (updates merge by key).",
        "file        — lib/i18n.ts (composer overlay) or lib/app-i18n (app UI).",
        "screen      — UI section where the key first appeared.",
        "",
        "Hero eyebrow on the home overview is privateDiveArchive (not",
        "diveLogCompanion). diveLogCompanion is the top-bar brand subtitle.",
        "",
        "When done, give the updated xlsx back with the prompt in chat.",
        "Keep the sheet name “All (screen order)” or say which sheet to import.",
    ]
    for i, line in enumerate(lines, start=2):
        readme[f"A{i}"] = line
    readme.column_dimensions["A"].width = 88

    # Hero example
    hero = wb.create_sheet("Home hero (example)")
    hero.append(["key", "EN", "JA", "ZH", "screenOrder", "fileOrder", "file", "screen"])
    for col in range(1, 9):
        cell = hero.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
    for i, key in enumerate(
        ("privateDiveArchive", "heroTitle", "heroDescription"), start=1
    ):
        hero.append(
            [
                key,
                en.get(key, ""),
                ja.get(key, ""),
                zh.get(key, ""),
                i,
                file_order.get(key, ""),
                key_file.get(key, ""),
                "01 Home — top bar & overview",
            ]
        )
    for col, width in zip("ABCDEFGH", (28, 50, 50, 50, 12, 10, 14, 28)):
        hero.column_dimensions[col].width = width

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Data rows: {data_rows}")


if __name__ == "__main__":
    main()
