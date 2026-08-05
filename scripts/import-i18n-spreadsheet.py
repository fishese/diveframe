"""Import EN/JA/ZH strings from i18n-list.xlsx back into source files.

Default sheet: "All (screen order)"
Match by key. Skip section-header rows (key starts with "—").
Do not rename keys or reorder TypeScript object keys.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path(r"C:\Users\Selena\OneDrive\Documents\i18n-list.xlsx")
DEFAULT_SHEET = "All (screen order)"

PAIR_RE = re.compile(
    r'^  ([A-Za-z][A-Za-z0-9_]*)\s*:\s*("(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\')',
    re.M,
)


def unescape_js_string(raw: str) -> str:
    body = raw[1:-1]
    out: list[str] = []
    i = 0
    while i < len(body):
        if body[i] == "\\" and i + 1 < len(body):
            nxt = body[i + 1]
            mapping = {"n": "\n", "t": "\t", "\\": "\\", '"': '"', "'": "'"}
            out.append(mapping.get(nxt, nxt))
            i += 2
            continue
        out.append(body[i])
        i += 1
    return "".join(out)


def escape_js_string(value: str, quote: str) -> str:
    """Escape a JS string body to match the file's quote style."""
    out: list[str] = []
    for ch in value:
        if ch == "\\":
            out.append("\\\\")
        elif ch == "\n":
            out.append("\\n")
        elif ch == "\t":
            out.append("\\t")
        elif ch == quote:
            out.append("\\" + quote)
        else:
            out.append(ch)
    return "".join(out)


def parse_object_block(text: str, marker: str) -> dict[str, str]:
    start = text.find(marker)
    if start < 0:
        return {}
    brace = text.find("{", start)
    if brace < 0:
        return {}
    depth = 0
    end = brace
    for i, ch in enumerate(text[brace:], start=brace):
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                end = i
                break
    block = text[brace : end + 1]
    return {m.group(1): unescape_js_string(m.group(2)) for m in PAIR_RE.finditer(block)}


def parse_locale_ordered(path: Path) -> tuple[dict[str, str], list[str]]:
    text = path.read_text(encoding="utf-8")
    ordered = [m.group(1) for m in PAIR_RE.finditer(text)]
    values = {m.group(1): unescape_js_string(m.group(2)) for m in PAIR_RE.finditer(text)}
    return values, ordered


def load_repo() -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, str],
    dict[str, str],
]:
    """Returns en, ja, zh, key->file."""
    overlay = (ROOT / "lib" / "i18n.ts").read_text(encoding="utf-8")
    overlay_en = parse_object_block(overlay, "en:")
    overlay_ja = parse_object_block(overlay, "ja:")
    overlay_zh = parse_object_block(overlay, "zh-Hant:")
    app_en, _ = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "en.ts")
    app_ja, _ = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "ja.ts")
    app_zh, _ = parse_locale_ordered(ROOT / "lib" / "app-i18n" / "zh-Hant.ts")

    en = {**overlay_en, **app_en}
    ja = {**overlay_ja, **app_ja}
    zh = {**overlay_zh, **app_zh}

    key_file: dict[str, str] = {}
    for key in overlay_en:
        key_file[key] = "lib/i18n.ts"
    for key in app_en:
        key_file[key] = "lib/app-i18n"
    for key in set(en) | set(ja) | set(zh):
        if key not in key_file:
            key_file[key] = "lib/app-i18n"
    return en, ja, zh, key_file


def load_sheet(
    xlsx: Path, sheet_name: str
) -> list[dict[str, str]]:
    wb = load_workbook(xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    sheet = wb[sheet_name]
    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    col = {h: i for i, h in enumerate(headers)}
    for required in ("key", "EN", "JA", "ZH"):
        if required not in col:
            raise SystemExit(f"Missing column {required!r} in sheet headers: {headers}")

    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key_raw = row[col["key"]]
        if key_raw is None or str(key_raw).strip() == "":
            continue
        key = str(key_raw).strip()
        if key.startswith("—") or key.startswith("–"):
            continue
        file_val = ""
        if "file" in col and row[col["file"]] is not None:
            file_val = str(row[col["file"]]).strip()
        rows.append(
            {
                "key": key,
                "EN": "" if row[col["EN"]] is None else str(row[col["EN"]]),
                "JA": "" if row[col["JA"]] is None else str(row[col["JA"]]),
                "ZH": "" if row[col["ZH"]] is None else str(row[col["ZH"]]),
                "file": file_val,
            }
        )
    return rows


def replace_key_value(text: str, key: str, new_value: str) -> tuple[str, bool]:
    """Replace a single key's string value, preserving quote style. Returns (text, changed)."""
    pattern = re.compile(
        rf'^(  {re.escape(key)}\s*:\s*)("(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\')',
        re.M,
    )
    match = pattern.search(text)
    if not match:
        return text, False
    raw = match.group(2)
    quote = raw[0]
    old_value = unescape_js_string(raw)
    if old_value == new_value:
        return text, False
    escaped = escape_js_string(new_value, quote)
    replacement = f"{match.group(1)}{quote}{escaped}{quote}"
    return text[: match.start()] + replacement + text[match.end() :], True


def apply_overlay(
    path: Path,
    updates: dict[str, dict[str, str]],
) -> dict[str, int]:
    """Apply updates to lib/i18n.ts locale blocks. updates[key] = {EN,JA,ZH}."""
    text = path.read_text(encoding="utf-8")
    counts = {"EN": 0, "JA": 0, "ZH": 0}
    markers = [("EN", "en:"), ("JA", "ja:"), ("ZH", "zh-Hant:")]

    for loc, marker in markers:
        start = text.find(marker)
        if start < 0:
            continue
        brace = text.find("{", start)
        depth = 0
        end = brace
        for i, ch in enumerate(text[brace:], start=brace):
            if ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = i
                    break
        block = text[brace : end + 1]
        new_block = block
        for key, vals in updates.items():
            new_block, changed = replace_key_value(new_block, key, vals[loc])
            if changed:
                counts[loc] += 1
        text = text[:brace] + new_block + text[end + 1 :]

    path.write_text(text, encoding="utf-8")
    return counts


def apply_locale_file(path: Path, updates: dict[str, str]) -> int:
    text = path.read_text(encoding="utf-8")
    changed = 0
    for key, value in updates.items():
        text, did = replace_key_value(text, key, value)
        if did:
            changed += 1
    path.write_text(text, encoding="utf-8")
    return changed


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Write changes (default is dry-run / mismatch report only)",
    )
    parser.add_argument(
        "--force-mismatches",
        action="store_true",
        help="Apply even if sheet/repo key sets differ (still skip unknown sheet keys)",
    )
    args = parser.parse_args()

    en, ja, zh, key_file = load_repo()
    repo_keys = set(en) | set(ja) | set(zh)
    rows = load_sheet(args.xlsx, args.sheet)
    sheet_keys = {r["key"] for r in rows}

    only_sheet = sorted(sheet_keys - repo_keys)
    only_repo = sorted(repo_keys - sheet_keys)

    print(f"Sheet: {args.sheet!r} from {args.xlsx}")
    print(f"Sheet data rows: {len(rows)}")
    print(f"Repo keys: {len(repo_keys)}")
    print()
    print(f"Keys only in sheet ({len(only_sheet)}):")
    for k in only_sheet:
        print(f"  + {k}")
    if not only_sheet:
        print("  (none)")
    print()
    print(f"Keys only in repo ({len(only_repo)}):")
    for k in only_repo:
        print(f"  - {k}")
    if not only_repo:
        print("  (none)")
    print()

    # Planned changes among matching keys
    planned: dict[str, dict[str, dict[str, str]]] = {
        "lib/i18n.ts": {},
        "lib/app-i18n": {},
    }
    summary: dict[str, dict[str, int]] = {
        "lib/i18n.ts": {"EN": 0, "JA": 0, "ZH": 0},
        "lib/app-i18n/en.ts": {"EN": 0},
        "lib/app-i18n/ja.ts": {"JA": 0},
        "lib/app-i18n/zh-Hant.ts": {"ZH": 0},
    }

    for r in rows:
        key = r["key"]
        if key not in repo_keys:
            continue
        target = r["file"] or key_file.get(key, "lib/app-i18n")
        if target not in ("lib/i18n.ts", "lib/app-i18n"):
            # Normalize common variants
            if "i18n.ts" in target and "app-i18n" not in target:
                target = "lib/i18n.ts"
            else:
                target = "lib/app-i18n"

        diffs: dict[str, str] = {}
        if r["EN"] != en.get(key, ""):
            diffs["EN"] = r["EN"]
        if r["JA"] != ja.get(key, ""):
            diffs["JA"] = r["JA"]
        if r["ZH"] != zh.get(key, ""):
            diffs["ZH"] = r["ZH"]
        if not diffs:
            continue

        planned[target][key] = {
            "EN": r["EN"],
            "JA": r["JA"],
            "ZH": r["ZH"],
        }
        if target == "lib/i18n.ts":
            for loc in diffs:
                summary["lib/i18n.ts"][loc] += 1
        else:
            if "EN" in diffs:
                summary["lib/app-i18n/en.ts"]["EN"] += 1
            if "JA" in diffs:
                summary["lib/app-i18n/ja.ts"]["JA"] += 1
            if "ZH" in diffs:
                summary["lib/app-i18n/zh-Hant.ts"]["ZH"] += 1

    total_keys = len(planned["lib/i18n.ts"]) + len(planned["lib/app-i18n"])
    print(f"Keys with at least one locale change: {total_keys}")
    print("Planned change counts:")
    for path, locs in summary.items():
        parts = ", ".join(f"{loc}={n}" for loc, n in locs.items() if n)
        if parts:
            print(f"  {path}: {parts}")
        else:
            print(f"  {path}: (none)")

    if only_sheet or only_repo:
        if not args.apply:
            print()
            print(
                "Mismatches found. Not writing. "
                "Re-run with --apply --force-mismatches to write matching keys only."
            )
            sys.exit(2)
        if not args.force_mismatches:
            print()
            print(
                "Mismatches found and --force-mismatches not set. Aborting write."
            )
            sys.exit(2)

    if not args.apply:
        print()
        print("Dry-run only. Re-run with --apply to write.")
        return

    # Apply
    overlay_updates = planned["lib/i18n.ts"]
    if overlay_updates:
        apply_overlay(ROOT / "lib" / "i18n.ts", overlay_updates)

    app_updates = planned["lib/app-i18n"]
    if app_updates:
        apply_locale_file(
            ROOT / "lib" / "app-i18n" / "en.ts",
            {k: v["EN"] for k, v in app_updates.items()},
        )
        apply_locale_file(
            ROOT / "lib" / "app-i18n" / "ja.ts",
            {k: v["JA"] for k, v in app_updates.items()},
        )
        apply_locale_file(
            ROOT / "lib" / "app-i18n" / "zh-Hant.ts",
            {k: v["ZH"] for k, v in app_updates.items()},
        )

    print()
    print("Applied updates.")
    print("Change counts:")
    for path, locs in summary.items():
        parts = ", ".join(f"{loc}={n}" for loc, n in locs.items() if n)
        if parts:
            print(f"  {path}: {parts}")


if __name__ == "__main__":
    main()
