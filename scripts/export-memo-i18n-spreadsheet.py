"""Export memo / memo-match i18n keys to an easy-edit Excel workbook."""

from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parents[1]
OUT_DOCS = Path(r"C:\Users\Selena\OneDrive\Documents\memo-i18n.xlsx")
OUT_REPO = ROOT / "docs" / "memo-i18n.xlsx"

PAIR_RE = re.compile(
    r'^  ([A-Za-z][A-Za-z0-9_]*)\s*:\s*("(?:\\.|[^"\\])*"|\'(?:\\.|[^\'\\])*\')',
    re.M,
)

GROUPS: list[tuple[str, list[str]]] = [
    (
        "Memo page",
        [
            "diveMemosEyebrow",
            "diveMemosTitle",
            "diveMemosIntro",
            "diveMemosAdd",
            "diveMemosHeading",
            "diveMemosDate",
            "diveMemosTime",
            "diveMemosHour",
            "diveMemosHourUp",
            "diveMemosHourDown",
            "diveMemosMinute",
            "diveMemosMeridiem",
            "diveMemosLocation",
            "diveMemosCoordinates",
            "diveMemosNoCoordinates",
            "diveMemosUseGps",
            "diveMemosPhotoGps",
            "diveMemosClearGps",
            "deleteMemo",
            "memoGpsUnsupported",
            "memoGpsCaptured",
            "memoGpsFailed",
            "importGuideMemosPrompt",
            "openDiveMemos",
        ],
    ),
    (
        "Memo ↔ dive matching",
        [
            "memoMatchTitle",
            "memoMatchTitleFromMemo",
            "memoMatchShow12h",
            "memoMatchShow24h",
            "memoMatchApplyEmpty",
            "memoMatchNothingToApply",
            "memoMatchCopyLocation",
            "memoMatchCopyGps",
            "memoMatchCopyBuddies",
            "memoMatchCopyNotes",
            "memoMatchDeleteConfirm",
            "memoMatchAppliedTitle",
            "memoMatchKeepMemo",
            "memoMatchDeleteMemo",
            "memoMatchLinkedNote",
            "memoMatchNoCandidates",
        ],
    ),
    (
        "Shared labels",
        [
            "buddy",
            "notes",
        ],
    ),
]

NOTES = {
    "diveMemosEyebrow": "Defined but currently unused in UI",
    "memoMatchLinkedNote": "Defined but currently unused in UI; keep {number}",
    "buddy": "Shared dive/memo field label",
    "notes": "Shared dive/memo field label",
    "memoMatchDeleteConfirm": "Used for confirm() and delete button aria/title",
    "memoMatchDeleteMemo": "Post-apply dialog button; same question wording as confirm",
}


def unescape_js_string(raw: str) -> str:
    body = raw[1:-1]
    out: list[str] = []
    i = 0
    while i < len(body):
        if body[i] == "\\" and i + 1 < len(body):
            nxt = body[i + 1]
            mapping = {"n": "\n", "t": "\t", "\\": "\\", '"': '"', "'": "'"}
            out.append(mapping.get(nxt, nxt))
            i += 2
            continue
        out.append(body[i])
        i += 1
    return "".join(out)


def parse_locale(path: Path) -> dict[str, str]:
    text = path.read_text(encoding="utf-8")
    return {m.group(1): unescape_js_string(m.group(2)) for m in PAIR_RE.finditer(text)}


def main() -> None:
    en = parse_locale(ROOT / "lib" / "app-i18n" / "en.ts")
    ja = parse_locale(ROOT / "lib" / "app-i18n" / "ja.ts")
    zh = parse_locale(ROOT / "lib" / "app-i18n" / "zh-Hant.ts")

    wb = Workbook()
    ws = wb.active
    ws.title = "Memo i18n"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D2731")
    helper_fill = PatternFill("solid", fgColor="1A3A44")
    section_fill = PatternFill("solid", fgColor="143946")
    section_font = Font(bold=True, color="8DEBD7")
    unused_fill = PatternFill("solid", fgColor="FFF4CE")
    thin = Border(
        left=Side(style="thin", color="2A4A55"),
        right=Side(style="thin", color="2A4A55"),
        top=Side(style="thin", color="2A4A55"),
        bottom=Side(style="thin", color="2A4A55"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = ["key", "EN", "JA", "ZH", "section", "placeholders", "notes"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = helper_fill if col >= 5 else header_fill
        cell.border = thin

    missing: list[str] = []
    key_count = 0
    for section, keys in GROUPS:
        ws.append([f"— {section}", "", "", "", section, "", ""])
        row = ws.max_row
        for col in range(1, 8):
            cell = ws.cell(row, col)
            cell.fill = section_fill
            cell.font = section_font
            cell.border = thin

        for key in keys:
            if key not in en:
                missing.append(key)
            en_v = en.get(key, "")
            ja_v = ja.get(key, "")
            zh_v = zh.get(key, "")
            note = NOTES.get(key, "")
            placeholders = " ".join(re.findall(r"\{[A-Za-z0-9_]+\}", en_v))
            ws.append([key, en_v, ja_v, zh_v, section, placeholders, note])
            key_count += 1
            row = ws.max_row
            for col in range(1, 8):
                cell = ws.cell(row, col)
                cell.alignment = wrap
                cell.border = thin
                if note.startswith("Defined but currently unused"):
                    cell.fill = unused_fill

    widths = {"A": 28, "B": 55, "C": 55, "D": 55, "E": 22, "F": 16, "G": 52}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"
    ws.row_dimensions[1].height = 22

    info = wb.create_sheet("README", 0)
    info["A1"] = "DiveFrame memo / memo-match i18n"
    info["A1"].font = Font(bold=True, size=14)
    info["A3"] = "Edit EN / JA / ZH columns. Do not rename keys."
    info["A4"] = "Yellow rows: keys defined in source but currently unused in UI."
    info["A5"] = "Keep placeholders like {number} intact in all languages."
    info["A6"] = "Source files: lib/app-i18n/en.ts, ja.ts, zh-Hant.ts"
    info["A7"] = "After editing, ask to import the changes back into the repo."
    info["A8"] = f"Generated from: {ROOT}"
    info.column_dimensions["A"].width = 90

    OUT_DOCS.parent.mkdir(parents=True, exist_ok=True)
    OUT_REPO.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_DOCS)
    wb.save(OUT_REPO)

    print(f"saved: {OUT_DOCS}")
    print(f"saved: {OUT_REPO}")
    print(f"keys: {key_count}")
    if missing:
        raise SystemExit(f"MISSING keys: {missing}")


if __name__ == "__main__":
    main()
